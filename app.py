from flask import Flask, render_template, request, redirect
from openpyxl import Workbook, load_workbook
import os

app = Flask(__name__)
FILE_NAME = "data.xlsx"

# Create Excel file if not exists
if not os.path.exists(FILE_NAME):
    wb = Workbook()
    ws = wb.active
    ws.append([
        "Name",
        "Emp ID",
        "Topics",
        "Marks (Clarity)",
        "Marks (Structure)",
        "Marks (Engagement)",
        "Date & Time",
    ])  # Excel headers
    wb.save(FILE_NAME)

@app.route("/", methods=["GET", "POST"])
def form():
    if request.method == "POST":
        # Use .get() to avoid BadRequestKeyError when keys are missing
        name = request.form.get("name", "").strip()
        emp_id = request.form.get("emp_id", "").strip()
        topics = request.form.get("topics", "").strip()

        def safe_int(val):
            try:
                return int(val)
            except (ValueError, TypeError):
                return None

        marks_clarity = safe_int(request.form.get("marks_clarity", ""))
        marks_structure = safe_int(request.form.get("marks_structure", ""))
        marks_engagement = safe_int(request.form.get("marks_engagement", ""))

        datetime_val = request.form.get("datetime", "").strip()

        # Save to Excel
        wb = load_workbook(FILE_NAME)
        ws = wb.active
        ws.append([
            name,
            emp_id,
            topics,
            marks_clarity if marks_clarity is not None else "",
            marks_structure if marks_structure is not None else "",
            marks_engagement if marks_engagement is not None else "",
            datetime_val,
        ])
        wb.save(FILE_NAME)

        return redirect("/")   # refresh form

    return render_template("form.html")

if __name__ == "__main__":
    # Bind to PORT environment variable (used by hosts like Render)
    port = int(os.environ.get("PORT", 5000))
    # If running on a hosting platform that provides PORT (e.g. Render),
    # force binding to 0.0.0.0 and disable the Werkzeug reloader so the
    # process listens on all interfaces and the host can detect the port.
    is_hosted = "PORT" in os.environ
    host = os.environ.get("HOST", "0.0.0.0") if not is_hosted else "0.0.0.0"
    # Allow enabling debug via FLASK_DEBUG env var, but avoid the reloader on hosts
    debug = os.environ.get("FLASK_DEBUG", "false").lower() in ("1", "true", "yes")
    if is_hosted:
        app.run(host=host, port=port, debug=debug, use_reloader=False)
    else:
        app.run(host=host, port=port, debug=debug)


